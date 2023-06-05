import time
import os
import pickle
import re

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.webdriver import WebDriver
from loguru import logger
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from dotenv import load_dotenv, find_dotenv
from bs4 import BeautifulSoup
from multiprocessing import Pool
from typing import Any


headers = {
    "profession": "director",
    "sortOrder": "STARMETER_DESC",
    "BeginYear": "2020",
    "EndYear": "2023"
}

# url with no years options
# "https://pro.imdb.com/discover/people?profession=director&ref_=hm_nv_ppl_dir&minNumOfReleasedCredits=1&sortOrder=STARMETER_DESC&pageNumber=1&creditEndYear=2023&creditBeginYear=2020&includeDeceased=true"


def create_excel() -> None:
    """
    Creating 2 Excel files with Name and Email columns title.
    :return: None
    """
    wb_direct = Workbook()
    wb_other = Workbook()

    ws_direct = wb_direct.active
    ws_other = wb_other.active

    ws_direct.title = headers['profession']
    ws_other.title = headers['profession']

    i_row_direct = 1
    i_row_other = 1
    ws_direct[f'A{i_row_direct}'] = 'Name'
    ws_direct[f'B{i_row_direct}'] = 'Email'
    ws_other[f'A{i_row_other}'] = 'Name'
    ws_other[f'B{i_row_other}'] = 'Email'

    wb_direct.save(f"{headers['profession']}-direct.xlsx")
    wb_other.save(f"{headers['profession']}-managers&agents.xlsx")


def write_data_excel(
        direct_contacts_list,
        other_contacts_list) -> None:  # list[tuple[Any, str] | tuple[Any, Any]]
    """
    This function write each person's email and name and surname in Excel rows.
    :param direct_contacts_list: list of people who has direct email
    :param other_contacts_list: managers or agents list who also could have an email
    :return: None
    """

    wb_direct = load_workbook(f"{headers['profession']}-direct.xlsx")
    wb_other = load_workbook(f"{headers['profession']}-managers&agents.xlsx")

    ws = wb_direct[f"{headers['profession']}"]
    i_row = ws.max_row + 1
    # logger.debug(i_row)
    for i_contact in direct_contacts_list:
        ws[f'A{i_row}'] = i_contact[0]
        ws[f'B{i_row}'] = i_contact[1]
        i_row += 1

    ws = wb_other[f"{headers['profession']}"]
    i_row = ws.max_row + 1
    # logger.debug(i_row)
    for i_contact in other_contacts_list:
        ws[f'A{i_row}'] = i_contact[0]
        ws[f'B{i_row}'] = i_contact[1]
        i_row += 1

    wb_direct.save(f"{headers['profession']}-direct.xlsx")
    wb_other.save(f"{headers['profession']}-managers&agents.xlsx")


def open_page_with_cookies(start_page_url: str, people_list_url: str, driver: WebDriver) -> str:
    """
    If there are cookies already in work directory, this function could help to log in with it.
    :param start_page_url: https://pro.imdb.com
    :param people_list_url: url with chosen sort parameters and profession search
    :param driver: WebDriver item
    :return: html code of page with people list
    """
    driver.get(url=start_page_url)

    for cookie in pickle.load(open("fox_cookies", "rb")):
        driver.add_cookie(cookie)

    driver.refresh()
    driver.get(url=people_list_url)
    time.sleep(15)

    return driver.page_source


def getting_scrap_data(html: str, driver: WebDriver, people_list_url: str) -> None:
    """
    Getting all contacts and names of people, who signed up on pro.imdb.com
    :param html: page code
    :param driver: WebDriver item
    :param people_list_url: url with chosen sort parameters and profession search
    :return: None
    """
    soup = BeautifulSoup(html, 'html.parser')
    # each_person_in_list - list with each person html segment on page, len() is 50 to 0.
    each_person_in_list = soup.find_all('div', class_='a-section a-spacing-none a-padding-base aok-relative')
    each_contacts_segment = soup.find_all('div', class_='a-section a-spacing-none search-results__contacts')
    direct_contacts = list()
    other_contacts = list()

    for i in range(len(each_person_in_list)):
        # searching any contact info. If there is no contacts, set None. So, either contacts_info_list is None
        # or some html in span tag.
        # contacts_info_list = each_contacts_segment[i].find('span', class_='a-truncate-cut')
        try:
            contacts_info = each_contacts_segment[i].find('span', class_='a-truncate-cut')
            # logger.debug(f'{contacts_info.text}')
            if 'Direct Contact' in contacts_info.text:
                personal_account_link = ''.join(
                    ['https://pro.imdb.com', each_person_in_list[i].find('a', class_='a-link-normal').get('href')]
                )
                try:
                    # searching email address in contacts html segment
                    data_contacts = contacts_info.find('br', recursive=True)
                    if data_contacts and '@' not in data_contacts.next_sibling:
                        driver.get(url=personal_account_link)
                        time.sleep(7)
                        try:
                            email = driver.find_element(
                                By.XPATH,
                                "//span[contains(@class, 'a-list-item')]/a[contains(@class, 'a-size- a-align- a-link- clickable_share_link')]"
                            ).text
                            direct_contacts.append((contacts_info.find('br').previous_sibling[2:], email))
                        except Exception:
                            pass
                    elif data_contacts:
                        direct_contacts.append((
                            contacts_info.find('br').previous_sibling[2:],
                            contacts_info.find('br', recursive=True).next_sibling
                        ))
                    else:
                        driver.get(url=personal_account_link)
                        time.sleep(7)
                        try:
                            email = driver.find_element(
                                By.XPATH,
                                "//span[contains(@class, 'a-list-item')]/a[contains(@class, 'a-size- a-align- a-link- clickable_share_link')]"
                            ).text
                            direct_contacts.append((contacts_info.text.replace('Direct Contact: ', ''), email))
                        except Exception:
                            pass
                except Exception:  # if there is no email in contact info on page with people list
                    pass
            else:
                # if there is only manager or agent contact
                contacts_info = contacts_info.find(
                    'a', 'a-color-base a-link-normal search-results__not-toggle-card')
                # url is personal account link like 'https://pro.imdb.com/name/nm2185453?ref_=dsc_pe_res_nm_view_7'
                driver.get(url=''.join(['https://pro.imdb.com', contacts_info.get('href')]))
                try:
                    email = driver.find_element(
                        By.XPATH,
                        "//span[contains(@class, 'a-list-item')]/a[contains(@class, 'a-size- a-align- a-link- clickable_share_link')]"
                    ).text
                    other_contacts.append((contacts_info.text, email))
                except Exception:
                    # logger.debug("No email info on someone's page.")
                    pass
        except Exception:
            # logger.debug('Empty contact info.')
            pass

    # logger.info(re.search(r'pageNumber=(\d+)', pages_with_optional_url).group(1))
    logger.debug(re.search(r'pageNumber=(\d+)', people_list_url).group(1))
    logger.debug(f"\ndirect contacts - {direct_contacts}\nother - {other_contacts}")
    write_data_excel(direct_contacts_list=direct_contacts, other_contacts_list=other_contacts)
    # end = time.time() - start
    # logger.debug(end)


def scrap_func(pages_with_optional_url: str) -> None:
    """
    Creating webdriver items and set some default parameters for searching.
    :param pages_with_optional_url: url with chosen sort parameters and profession search
    :return: None
    """
    # .env file check
    if not find_dotenv():
        exit("Environment variables are not loaded because there is no .env file")
    else:
        load_dotenv()

    # user-agent activation
    useragent = UserAgent()

    # options
    options = Options()
    options.set_preference("general.useragent.override", useragent.random)

    # headless mode
    options.add_argument("--headless")
    # options.headless = True

    # disable webdriver mode
    options.set_preference("dom.webdriver.enabled", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    # present(failed)

    # IMDb urls
    login_url = "https://secure.imdb.com/ap/signin?openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.imdb.com%2Fregistration%2Fap-signin-handler%2Fimdb_pro_us&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=imdb_pro_us&openid.mode=checkid_setup&siteState=eyJvcGVuaWQuYXNzb2NfaGFuZGxlIjoiaW1kYl9wcm9fdXMiLCJyZWRpcmVjdFRvIjoiaHR0cHM6Ly9wcm8uaW1kYi5jb20vbG9naW4vbHdhIn0&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0"
    origin_url = "https://pro.imdb.com"

    # driver activation
    webdriver_service = Service('/IMDb_scraper/firefoxdirectory/geckodriver')
    driver = webdriver.Firefox(service=webdriver_service, options=options)

    try:
        # # to see if webdriver is disabled
        # driver.get("https://intoli.com/blog/not-possible-to-block-chrome-headless/chrome-headless-test.html")

        # # to get cookies
        # driver.get(url=login_url)

        # # enter your data to log in to get cookies
        # email_input = driver.find_element(By.ID, 'ap_email')
        # # if some text has already been entered in the field
        # email_input.clear()
        # email_input.send_keys(os.getenv("EMAIL"))
        # password_input = driver.find_element(By.ID, 'ap_password')
        # password_input.send_keys(os.getenv("PASSWORD"))
        # password_input.send_keys(Keys.ENTER)
        # tap_login_button = driver.find_element(By.ID, 'signInSubmit').click()

        # driver.get(url=origin_url)  # getting start page which is https://pro.imdb.com

        # cookies
        # pickle.dump(driver.get_cookies(), open("fox_cookies", "wb"))  # write cookies
        # for cookie in pickle.load(open("fox_cookies", "rb")):  # getting cookies for log in
        #     driver.add_cookie(cookie)

        # driver.refresh()
        # driver.get(url=pages_with_optional_url)  # getting people list page
        # time.sleep(15)

        # # getting pages number
        # pages_num = driver.find_elements(By.CLASS_NAME, 'a-disabled')[-1].text
        # logger.debug(pages_num)

        # html_content = driver.page_source  # getting html from people list page
        html_content = open_page_with_cookies(
            start_page_url=origin_url, people_list_url=pages_with_optional_url, driver=driver
        )
        # # Write the parsed HTML to a file with proper indentation
        # with open('IDMb_authorize.html', 'w') as file:
        #     file.write(html_content)

        # start = time.time()

        if "D'oh!" in html_content:  # if the server thought we were a ddos attack
            time.sleep(16 * 60)
            html_content = open_page_with_cookies(
                start_page_url=origin_url, people_list_url=pages_with_optional_url, driver=driver
            )
            getting_scrap_data(html=html_content, driver=driver, people_list_url=pages_with_optional_url)
        else:
            getting_scrap_data(html=html_content, driver=driver, people_list_url=pages_with_optional_url)

    except Exception as ex:
        logger.debug(ex)
    finally:
        driver.close()
        driver.quit()


if __name__ == "__main__":
    create_excel()
    pages_number = 3240
    simultaneously_work_pages_num = 6
    iter_num = pages_number // simultaneously_work_pages_num \
        if pages_number % simultaneously_work_pages_num == 0 \
        else pages_number // simultaneously_work_pages_num + 1
    for i_page in range(1, iter_num, simultaneously_work_pages_num):
        url_list = [
            f"https://pro.imdb.com/discover/people?profession={headers['profession']}&ref_=hm_nv_ppl_dir&minNumOfReleasedCredits=1&sortOrder={headers['sortOrder']}&pageNumber={page_number}&creditEndYear={headers['EndYear']}&creditBeginYear={headers['BeginYear']}&includeDeceased=true"
            for page_number in range(i_page, i_page + simultaneously_work_pages_num)
        ]
        p = Pool(processes=len(url_list))
        p.map(scrap_func, url_list)
        time.sleep(15)
